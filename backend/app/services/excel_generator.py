"""
Excel export generator using openpyxl.
Creates a 5-sheet workbook: Comparison + one per company.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from io import BytesIO


# Indian currency format for Excel
INDIAN_CURRENCY_FMT = '[$₹-4000]#,##,##0.00'

# Company colors (header tints)
COMPANY_COLORS = {
    "Apollo": "FFDDD3",     # warm red tint
    "Supreme": "D3E4FD",    # blue tint
    "Astral": "D3F9D8",     # green tint
    "Ashirvad": "FEF3C7",   # amber tint
}

HEADER_FILL = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
BEST_PRICE_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=12)
GST_FONT = Font(name="Calibri", italic=True, size=11, color="555555")
PAYABLE_FONT = Font(name="Calibri", bold=True, size=13)
PAYABLE_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
)


def auto_width(ws, min_width=10, max_width=40):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def generate_quote_excel(quote_data: dict, dealer_profile: dict) -> bytes:
    """
    Generate multi-sheet Excel workbook for a quote.

    Sheets:
        1. Comparison — all products, all 4 company totals side by side
        2-5. Apollo, Supreme, Astral, Ashirvad — per-company detail

    Returns: Excel file bytes
    """
    wb = Workbook()
    companies = ["Apollo", "Supreme", "Astral", "Ashirvad"]
    line_items = [li for li in quote_data.get("line_items", []) if not li.get("error")]
    company_totals = {ct["company"]: ct for ct in quote_data.get("company_totals", [])}

    b_name = dealer_profile.get('dealer_name', '')
    b_addr = dealer_profile.get('address', '')
    b_gst = dealer_profile.get('gst_number', '')
    b_mob = dealer_profile.get('mobile_number', '')
    c_name = quote_data.get('client_name', '')
    c_date = quote_data.get('created_at', '')[:10] if quote_data.get('created_at') else ''

    def write_headers(ws, max_col_letter="J", is_comparison=False, company=""):
        ws.merge_cells(f"A1:{max_col_letter}1")
        title = b_name if is_comparison else f"{b_name} ({company})"
        ws["A1"] = title
        ws["A1"].font = Font(name="Calibri", bold=True, size=16)

        ws.merge_cells(f"A2:{max_col_letter}2")
        ws["A2"] = b_addr
        ws["A2"].font = Font(name="Calibri", size=11)

        ws.merge_cells(f"A3:{max_col_letter}3")
        ws["A3"] = f"GSTIN: {b_gst}" if b_gst else "GSTIN: N/A"
        ws["A3"].font = Font(name="Calibri", size=11)

        ws.merge_cells(f"A4:{max_col_letter}4")
        ws["A4"] = f"Mobile: {b_mob}" if b_mob else "Mobile: N/A"
        ws["A4"].font = Font(name="Calibri", size=11)

        ws.merge_cells(f"A5:{max_col_letter}5")
        ws["A5"] = f"Client: {c_name} | Date: {c_date}"
        ws["A5"].font = Font(name="Calibri", bold=True, size=12)

    # ─── Sheet 1: Comparison ───
    ws = wb.active
    ws.title = "Comparison"

    write_headers(ws, max_col_letter="J", is_comparison=True)

    # Headers
    headers = ["Sr.", "Item Code", "Description", "Category", "Material", "Qty",
               "Apollo Total", "Supreme Total", "Astral Total", "Ashirvad Total"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, item in enumerate(line_items, 8):
        ws.cell(row=row_idx, column=1, value=item["sr"])
        ws.cell(row=row_idx, column=2, value=item["product_code"])
        ws.cell(row=row_idx, column=3, value=item["description"])
        ws.cell(row=row_idx, column=4, value=item["category"])
        ws.cell(row=row_idx, column=5, value=item["material_type"])
        ws.cell(row=row_idx, column=6, value=float(item["quantity"]) if item.get("quantity") else 0)

        for comp_idx, company in enumerate(companies, 7):
            c_lower = company.lower()
            val = item.get(f"line_total_{c_lower}")
            cell = ws.cell(row=row_idx, column=comp_idx,
                          value=float(val) if val else 0)
            cell.number_format = INDIAN_CURRENCY_FMT

        # Alternating row colors
        if row_idx % 2 == 0:
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).fill = ALT_ROW_FILL

    # Grand Total row (subtotal before tax)
    total_row = 8 + len(line_items)
    ws.cell(row=total_row, column=1, value="").font = TOTAL_FONT
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    ws.cell(row=total_row, column=1, value="SUBTOTAL").font = TOTAL_FONT

    best_company = None
    for company in companies:
        ct = company_totals.get(company, {})
        if ct.get("is_best_price"):
            best_company = company

    for comp_idx, company in enumerate(companies, 7):
        ct = company_totals.get(company, {})
        total_val = ct.get("total", 0)
        cell = ws.cell(row=total_row, column=comp_idx,
                      value=float(total_val) if total_val else 0)
        cell.number_format = INDIAN_CURRENCY_FMT
        cell.font = TOTAL_FONT

    # CGST row
    cgst_row = total_row + 1
    first_ct = next(iter(company_totals.values()), {})
    cgst_rate = first_ct.get("cgst_rate", 0)
    sgst_rate = first_ct.get("sgst_rate", 0)

    ws.merge_cells(start_row=cgst_row, start_column=1, end_row=cgst_row, end_column=6)
    ws.cell(row=cgst_row, column=1, value=f"CGST @ {cgst_rate}%").font = GST_FONT
    for comp_idx, company in enumerate(companies, 7):
        ct = company_totals.get(company, {})
        cgst_amt = ct.get("cgst_amount", 0)
        cell = ws.cell(row=cgst_row, column=comp_idx, value=float(cgst_amt) if cgst_amt else 0)
        cell.number_format = INDIAN_CURRENCY_FMT
        cell.font = GST_FONT

    # SGST row
    sgst_row = cgst_row + 1
    ws.merge_cells(start_row=sgst_row, start_column=1, end_row=sgst_row, end_column=6)
    ws.cell(row=sgst_row, column=1, value=f"SGST @ {sgst_rate}%").font = GST_FONT
    for comp_idx, company in enumerate(companies, 7):
        ct = company_totals.get(company, {})
        sgst_amt = ct.get("sgst_amount", 0)
        cell = ws.cell(row=sgst_row, column=comp_idx, value=float(sgst_amt) if sgst_amt else 0)
        cell.number_format = INDIAN_CURRENCY_FMT
        cell.font = GST_FONT

    # Total Payable row
    payable_row = sgst_row + 1
    ws.merge_cells(start_row=payable_row, start_column=1, end_row=payable_row, end_column=6)
    ws.cell(row=payable_row, column=1, value="TOTAL PAYABLE (incl. GST)").font = PAYABLE_FONT
    for comp_idx, company in enumerate(companies, 7):
        ct = company_totals.get(company, {})
        total_gst = ct.get("total_with_gst", 0)
        cell = ws.cell(row=payable_row, column=comp_idx, value=float(total_gst) if total_gst else 0)
        cell.number_format = INDIAN_CURRENCY_FMT
        cell.font = PAYABLE_FONT
        cell.fill = PAYABLE_FILL
        if company == best_company:
            cell.fill = BEST_PRICE_FILL

    auto_width(ws)
    ws.column_dimensions['A'].width = 5  # Shorten Sr. column width

    # ─── Sheets 2-5: Per-company detail ───
    for company in companies:
        ws = wb.create_sheet(title=company)
        c_lower = company.lower()

        write_headers(ws, max_col_letter="J", is_comparison=False, company=company)

        # Headers
        headers = ["Sr.", "Item Code", "Description", "Category", "Material",
                   "MRP", "Qty", "Unit Price", "Line Total"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_idx, value=header)
            cell.fill = PatternFill(
                start_color=COMPANY_COLORS.get(company, "1A1A1A"),
                end_color=COMPANY_COLORS.get(company, "1A1A1A"),
                fill_type="solid"
            )
            cell.font = Font(name="Calibri", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, item in enumerate(line_items, 8):
            ws.cell(row=row_idx, column=1, value=item["sr"])
            ws.cell(row=row_idx, column=2, value=item["product_code"]).font = Font(name="Consolas")
            ws.cell(row=row_idx, column=3, value=item["description"])
            ws.cell(row=row_idx, column=4, value=item["category"])
            ws.cell(row=row_idx, column=5, value=item["material_type"])

            mrp = item.get(f"mrp_{c_lower}")
            cell = ws.cell(row=row_idx, column=6, value=float(mrp) if mrp else 0)
            cell.number_format = INDIAN_CURRENCY_FMT

            ws.cell(row=row_idx, column=7, value=float(item["quantity"]) if item.get("quantity") else 0)

            unit_price = item.get(f"unit_price_{c_lower}")
            cell = ws.cell(row=row_idx, column=8, value=float(unit_price) if unit_price else 0)
            cell.number_format = INDIAN_CURRENCY_FMT

            line_total = item.get(f"line_total_{c_lower}")
            cell = ws.cell(row=row_idx, column=9, value=float(line_total) if line_total else 0)
            cell.number_format = INDIAN_CURRENCY_FMT

            # Alternating rows
            if row_idx % 2 == 0:
                for col in range(1, 10):
                    ws.cell(row=row_idx, column=col).fill = ALT_ROW_FILL

        # Grand Total (subtotal)
        total_row = 8 + len(line_items)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
        ws.cell(row=total_row, column=1, value="SUBTOTAL").font = TOTAL_FONT

        ct = company_totals.get(company, {})
        total_val = ct.get("total", 0)
        cell = ws.cell(row=total_row, column=9, value=float(total_val) if total_val else 0)
        cell.number_format = INDIAN_CURRENCY_FMT
        cell.font = TOTAL_FONT

        # CGST row
        cgst_row = total_row + 1
        cgst_rate_val = ct.get("cgst_rate", 0)
        sgst_rate_val = ct.get("sgst_rate", 0)

        ws.merge_cells(start_row=cgst_row, start_column=1, end_row=cgst_row, end_column=8)
        ws.cell(row=cgst_row, column=1, value=f"CGST @ {cgst_rate_val}%").font = GST_FONT
        cgst_amt = ct.get("cgst_amount", 0)
        cell = ws.cell(row=cgst_row, column=9, value=float(cgst_amt) if cgst_amt else 0)
        cell.number_format = INDIAN_CURRENCY_FMT
        cell.font = GST_FONT

        # SGST row
        sgst_row = cgst_row + 1
        ws.merge_cells(start_row=sgst_row, start_column=1, end_row=sgst_row, end_column=8)
        ws.cell(row=sgst_row, column=1, value=f"SGST @ {sgst_rate_val}%").font = GST_FONT
        sgst_amt = ct.get("sgst_amount", 0)
        cell = ws.cell(row=sgst_row, column=9, value=float(sgst_amt) if sgst_amt else 0)
        cell.number_format = INDIAN_CURRENCY_FMT
        cell.font = GST_FONT

        # Total Payable row
        payable_row = sgst_row + 1
        ws.merge_cells(start_row=payable_row, start_column=1, end_row=payable_row, end_column=8)
        ws.cell(row=payable_row, column=1, value="TOTAL PAYABLE (incl. GST)").font = PAYABLE_FONT
        total_gst = ct.get("total_with_gst", 0)
        cell = ws.cell(row=payable_row, column=9, value=float(total_gst) if total_gst else 0)
        cell.number_format = INDIAN_CURRENCY_FMT
        cell.font = PAYABLE_FONT
        cell.fill = PAYABLE_FILL
        if ct.get("is_best_price"):
            cell.fill = BEST_PRICE_FILL

        auto_width(ws)
        ws.column_dimensions['A'].width = 5  # Shorten Sr. column width

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
